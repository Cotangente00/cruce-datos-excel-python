import tkinter as tk
from tkinter import filedialog, Label
from tkinter import messagebox
import os
import openpyxl
from manipular_INFORME_SOLICITUDES import *
from manipular_Hoja1 import *
import xlrd
import logging
import sys

def procesar_archivo_excel():
    filepath = filedialog.askopenfilename(title="Selecciona el archivo Excel a modificar", filetypes=[("Archivos Excel", "*.xlsx;*.xls")])
    if not filepath:
        return

    try:
        if filepath.endswith('.xlsx'):
            # Cargar el libro de Excel (xlsx)
            wb = openpyxl.load_workbook(filepath)
        elif filepath.endswith('.xls'):
            # Cargar el libro de Excel (xls)
            xls_workbook = xlrd.open_workbook(filepath)
            wb = openpyxl.Workbook()

            # Copiar hojas de xls a xlsx
            for sheet in xls_workbook.sheets():
                ws_xls = xls_workbook.sheet_by_name(sheet.name)
                ws_new = wb.create_sheet(sheet.name)

                for row in range(ws_xls.nrows):
                    for col in range(ws_xls.ncols):
                        cell_value = ws_xls.cell_value(row, col)
                        ws_new.cell(row=row+1, column=col+1, value=cell_value)

            # Eliminar la hoja predeterminada de Workbook
            wb.remove(wb.active)

        # Seleccionar la hoja de trabajo
        ws = wb.active

        # Mensaje de confirmación para que el usuario sea consciente que de el listado de expertas fue copiado en la celda A5
        confirmacion = messagebox.askyesno('Confirmar modificación', 'Asegurese de que el listado de expertas haya sido copiado en la celda A5 de la hoja "Hoja1", para evitar dañar el contenido del archivo. ¿Desea continuar?')
        if not confirmacion:
            return
        
        #buscar el marcador de modificación 
        marcador = ws['AZ1'] #Marcador en la celda AD1
        modificado_antes = marcador.value == 'MODIFICADO' 

        if modificado_antes:
            # Mensaje de confirmación si el archivo ha sido modificado por la aplicación previamente  
            resultado = messagebox.askyesno('Confirmar modificación', 'Este archivo ya ha sido modificado previamente, si continúa, el contenido del archivo será distorsionado. ¿Desea continuar?')
            if not resultado:
                return
        
        # Función compuesta de funciones que ejecutan los cambios necesarios en la hoja INFORME SOLICITUDES
        ejecucion_funciones(ws)

        #Volver a cargar las hojas del archivo después de ejecurtar los cambios en la hoja INFORME SOLICITUDES
        ws=wb['Hoja1']
        ws2=wb['INFORME SOLICITUDES']

        # Funciones que hacen los cambios en la Hoja1
        ejecucion_funciones2(ws,ws2)


        ws2['Q2'] = 'Expertas que NO tienen servicio' 

        #Agregar marcador de que el archivo ha sido modificado por la aplicación en la celdd AZ1        
        ws2['AZ1'] = 'MODIFICADO' 

        # Pedir al usuario la ruta y nombre para guardar el nuevo archivo
        filepath_save = filedialog.asksaveasfilename(title="Guardar archivo Excel modificado como", defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])
        if not filepath_save:
            return  # Salir si el usuario cancela el diálogo de guardado

        # Guardar el archivo modificado
        wb.save(filepath_save)
        messagebox.showinfo("Proceso completado", """
        • Cambios en la hoja "INFORME SOLICITUDES":
        - Se eliminaron las filas 1, 2, 3, y 4.
        - Total servicios, Tipo, Turno partido, Jornada fija, 
          Concepto: novedad y ausencias, Concepto: novedad 
          control empleados, CC experta cambios, Experta cambio,
          Notificación SMS, Notificación SMS cliente, SMS 
          enviado cliente. Columnas eliminadas.
        - Formato de fecha conservado DD/MM/YYYY.
        - Tamaño horizontal de las columnas adaptado.
        - Solicitud, Referencia Externa y Cédula modificada a 
          formato numérico.
        - Expertas que SI tienen novedad, resaltadas con color 
          amarillo.
        • Cambios en la hoja "Hoja1":
        - Nombres y apellidos concatendados.
        - Fecha, Sexo, localidad, número de celular y 
          TCVA eliminadas.
        - Datos trasladados a la celda D5.
        - BUSCARV desde Hoja1 a INFORME SOLICITUDES 
          número de documento y nombre completo en 
          las columnas M y N.
        - BUSCARV desde INFORME SOLICITUDES a Hoja1 
          nombre completo en la columna H.
        - Listado de expertas sin servicio copiado en las 
          columnas Q, R y S.
        """)
        abrir_excel(filepath_save)
    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error al procesar el archivo:\n{str(e)}")


# Configurar la interfaz gráfica
root = tk.Tk()
root.wm_title("Informe Solicitudes y Expertas Disponibles")
Label(root, text="Pegar todo el listado de expertas en la celda A5 de la hoja 'Hoja1'").pack(pady=10) 
root.geometry('420x110')
root.resizable(width=False, height=False)

#Función para que el ícono de la ventana funcione correctamente en cojunto con el comando 
def recurso_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


icon_path = recurso_path('icon.ico')
root.iconbitmap(icon_path)
btn_procesar_informe_solicitudes = tk.Button(root, text="Procesar Archivo Excel", command=procesar_archivo_excel).pack(pady=20)
root.mainloop() 