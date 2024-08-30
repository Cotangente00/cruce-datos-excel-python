import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import os


'''-------Sección para elminar filas innecesarias-------'''
def delete_filas(ws):
    #filas innecesarias a eliminar
    filas_eliminar = [1,2,3,4]

    # Iterar sobre las filas a eliminar
    for row_idx in sorted(filas_eliminar, reverse=True):
        # Eliminar la fila
        ws.delete_rows(row_idx) 


'''------Sección para eliminar columnas y ciudades innecesarias------'''
def delete_ciudades_columnas(ws):
    #columnas innecesarias a eliminar
    columnas_eliminar = ['C', 'D', 'J', 'K', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    # Eliminar las columnas
    for col in reversed(columnas_eliminar):
        # Convertir la letra de la columna a índice numérico
        col_idx = openpyxl.utils.cell.column_index_from_string(col)
        # Eliminar la columna utilizando el índice
        ws.delete_cols(col_idx)

    ciudades_permitidas = {'bogotá', 'cajica', 'chía', 'cota', 'soacha', 'bogota', 'cajicá', 'chia'}

    # Recorrer las filas en orden inverso para evitar problemas al eliminar filas
    for row in range(ws.max_row, 1, -1):
        ciudad = ws.cell(row=row, column=13).value.lower() # Obtener el valor de la Columna M (número 13) 
        if ciudad not in ciudades_permitidas and ciudad != '':
            ws.delete_rows(row, 1)
        elif ciudad == 'soacha': #Sin importar si el valor está en mayusculas o en minusculas
            ws.cell(row=row, column=17).value = 'Soacha(Validar servicio)' # Columna P, pero con la eliminación de la columna M, pasa a ser la columna O
        elif ciudad == '' or None:
            ws.cell(row=row, column=17).value = 'Ciudad vacía(Confirmar)' # Columna Q, pero con la eliminación de la columna M, pasa a ser la columna P


    # Columna a eliminar: ciudades una vez descartadas las que NO se necesitan
    columna_eliminar = 13

    # Elimina la columna M (ciudades) 
    ws.delete_cols(columna_eliminar, 1)  
    


'''------Sección para conservar el formato de fecha corta------'''
def date_format(ws):
    # Configurar el estilo de formato de fecha corta
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')

    # Aplicar el estilo a la columna de fechas (en este caso, columna D)
    for cell in ws['D']:  # Ajustar 'D' a la letra de la columna de fechas
        cell.style = date_style



'''------Sección para modificar el tamaño de las columnas y el aspecto------'''
def styles_columnSize(ws):
    ws.row_dimensions[1].height= 20

    fila = 1
    fila_excel = ws[f'A{fila}:Z{fila}']  # Establecemos un rango de columnas A a Z para la fila específica

    # Establecer el estilo de la fuente para negrita y subrayado
    estilo_negrita_subrayado = Font(bold=True, underline='single')

    # Aplicar el estilo a cada celda en la fila
    for fila_celdas in fila_excel:
        for celda in fila_celdas:
            celda.font = estilo_negrita_subrayado



'''------Sección para pasar todos lo números que están almacenados como texto a formato número------'''
def int_format(ws):
    columnas_a_convertir = ['A', 'B', 'J']

    # Iterar sobre las columnas a convertir
    for col in columnas_a_convertir:
        for cell in ws[col]:  # Iterar sobre todas las celdas de la columna
            # Convertir el valor de texto a número si es posible
            try:
                valor = float(cell.value)
                cell.value = valor  # Actualizar el valor en la celda con el número convertido
            except (ValueError, TypeError):
                # Manejar el caso donde no se puede convertir a número
                continue 



'''------Sección para resaltar a todas las que tienen novedades (nombre y cédula) y crear una hoja nueva------'''
def novedades_expertas(ws):
    # Color de fondo amarillo
    relleno_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Recorrer las filas (empezando desde la segunda fila por los encabezados)
    for fila in ws.iter_rows(min_row=2, min_col=10, max_col=11):  # Columnas N (14) y O (15)
        novedad = ws.cell(row=fila[0].row, column=13).value  # Columna M (13)

        # Verificar si la columna "novedad" es "Si"
        if novedad == "Si":
            # Aplicar formato de color amarillo a las celdas de nombre (N) y cedula (O)
            for celda in fila:
                celda.fill = relleno_amarillo

    # Columna a eliminar: Tiene_novedad una vez resaltadas las que SI tienen novedades
    columna_eliminar = 13

    # Elimina la columna R (Tiene novedad) 
    ws.delete_cols(columna_eliminar, 1)  


'''------Abrir archivo Excel automáticamente una vez hechos los cambios------'''
def abrir_excel(filepath):
    try:
        os.startfile(filepath)
    except OSError as e:
        print(f"No se pudo abrir el archivo '{filepath}': {e}")


'''------Función para ajustar el tamaño horizontal de las columnas de forma automática------'''

def ajustar_tamaño_columnas(ws):
    max_width = 200

    for column in ws.columns:
            max_length = 0
            column = column[0:]  
            for cell in column:
                try:  # Manejar posibles errores de tipo
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 4, max_width) if max_width else max_length + 4
            # Obtener la letra de la columna a partir del objeto columna
            column_letter = column[0].column_letter
            # Ajustar el ancho de la columna usando el nombre de la columna
            ws.column_dimensions[column_letter].width = adjusted_width

'''------Función que globaliza todas las funciones anteriores------'''
def ejecucion_funciones(ws):
    delete_filas(ws)
    delete_ciudades_columnas(ws)
    date_format(ws)
    styles_columnSize(ws)
    int_format(ws)