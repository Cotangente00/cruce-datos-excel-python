import openpyxl
from manipular_Hoja1 import *
from manipular_INFORME_SOLICITUDES import *
import xlwt



def convert_xlsx_to_xls(xlsx_file, xls_file):
    # Cargar el archivo .xlsx usando openpyxl
    wb_xlsx = openpyxl.load_workbook(xlsx_file)
    sheet_xlsx = wb_xlsx[wb_xlsx.sheetnames[1]]
    find_table_and_move_to_A5(sheet_xlsx)
   
    # Crear un nuevo archivo .xls usando xlwt
    wb_xls = xlwt.Workbook()

    # Iterar sobre todas las hojas del archivo .xlsx
    for sheet_name in wb_xlsx.sheetnames:
        sheet_xlsx = wb_xlsx[sheet_name]
        sheet_xls = wb_xls.add_sheet(sheet_name)

        # Copiar contenido de cada hoja del .xlsx a .xls
        for row_index, row in enumerate(sheet_xlsx.iter_rows()):
            for col_index, cell in enumerate(row):
                sheet_xls.write(row_index, col_index, cell.value)

    wb_xls.save(xls_file)


