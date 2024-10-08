import tkinter as tk
from tkinter import filedialog, Label
from tkinter import messagebox
import os
import openpyxl
from manipular_INFORME_SOLICITUDES import *
from manipular_Hoja1 import *
import xlrd
import logging
import sys
import os
import xlwt
from convert_xlsx_to_xls import *
from funciones_weekend import *
from funciones_botones import *   
import datetime



def intro_function_lunes_jueves(filepath):
    try:

        if filepath.endswith('.xlsx'):
            #convertir de xlsx a xls para convertirlo de vuelta
            convert_xlsx_to_xls(filepath, 'temp.xls')

            # Cargar el libro de Excel (xls)
            xls_workbook = xlrd.open_workbook('temp.xls')
            wb = openpyxl.Workbook()

            # Copiar hojas de xls a xlsx
            for sheet in xls_workbook.sheets():
                ws_xls = xls_workbook.sheet_by_name(sheet.name)
                ws_new = wb.create_sheet(sheet.name)

                for row in range(ws_xls.nrows):
                    for col in range(ws_xls.ncols):
                        cell_value = ws_xls.cell_value(row, col)
                        ws_new.cell(row=row+1, column=col+1, value=cell_value)
            
            wb.remove(wb.active)

            os.remove('temp.xls')

            ws = wb[wb.sheetnames[1]]
            a5 = ws['A5']
            if a5.value < 1000:
                ws.delete_rows(a5.row, 1)
                ws.delete_cols(a5.column, 1)

        elif filepath.endswith('.xls'):

            find_table_and_move_to_A5_xls(filepath, 'temp.xls')
            # Cargar el libro de Excel (xls)
            xls_workbook = xlrd.open_workbook('temp.xls')
            wb = openpyxl.Workbook()

            # Copiar hojas de xls a xlsx
            for sheet in xls_workbook.sheets():
                ws_xls = xls_workbook.sheet_by_name(sheet.name)
                ws_new = wb.create_sheet(sheet.name)

                for row in range(ws_xls.nrows):
                    for col in range(ws_xls.ncols):
                        cell_value = ws_xls.cell_value(row, col)
                        ws_new.cell(row=row+1, column=col+1, value=cell_value)

            # Eliminar la hoja predeterminada de Workbook
            wb.remove(wb.active)

            os.remove('temp.xls')
        
            ws = wb[wb.sheetnames[1]]
            a5 = ws['A5']
            if a5.value < 1000:
                ws.delete_rows(a5.row, 1)
                ws.delete_cols(a5.column, 1)

        button_lunes_jueves(wb)

    except Exception as e:
      messagebox.showerror("Error", f"Ha ocurrido un error al procesar el archivo:\n{str(e)}")

def intro_function_viernes_sabado(filepath):
    try:


        if filepath.endswith('.xlsx'):

            #convertir de xlsx a xls para convertirlo de vuelta
            convert_xlsx_to_xls_viernes_sabado(filepath, 'temp.xls')

            # Cargar el libro de Excel (xls)
            xls_workbook = xlrd.open_workbook('temp.xls')
            wb = openpyxl.Workbook()

            # Copiar hojas de xls a xlsx
            for sheet in xls_workbook.sheets():
                ws_xls = xls_workbook.sheet_by_name(sheet.name)
                ws_new = wb.create_sheet(sheet.name)

                for row in range(ws_xls.nrows):
                    for col in range(ws_xls.ncols):
                        cell_value = ws_xls.cell_value(row, col)
                        ws_new.cell(row=row+1, column=col+1, value=cell_value)
            
            wb.remove(wb.active)

            os.remove('temp.xls')
        
    
        elif filepath.endswith('.xls'):
      
            find_table_and_move_to_A5_xls_viernes_sabado(filepath, 'temp.xls')
            # Cargar el libro de Excel (xls)
            xls_workbook = xlrd.open_workbook('temp.xls')
            wb = openpyxl.Workbook()

            # Copiar hojas de xls a xlsx
            for sheet in xls_workbook.sheets():
                ws_xls = xls_workbook.sheet_by_name(sheet.name)
                ws_new = wb.create_sheet(sheet.name)

                for row in range(ws_xls.nrows):
                    for col in range(ws_xls.ncols):
                        cell_value = ws_xls.cell_value(row, col)
                        ws_new.cell(row=row+1, column=col+1, value=cell_value)

            # Eliminar la hoja predeterminada de Workbook
            wb.remove(wb.active)
      
            os.remove('temp.xls')

        # Seleccionar la hoja de trabajo
        button_viernes_sabado(wb)

    except Exception as e:
      messagebox.showerror("Error", f"Ha ocurrido un error al procesar el archivo:\n{str(e)}")