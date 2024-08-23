from openpyxl.utils.cell import get_column_letter
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl.styles import NamedStyle
from manipular_INFORME_SOLICITUDES import *


'''------Concatenar nombres y apellidos------'''
def concatenar_nombres_apellidos(ws):
    # Iterar sobre las filas
    for row in ws.iter_rows(min_row=5, min_col=2, max_col=3):  # empieza desde la quinta fila (min_row=5) A5
        # Obtener los valores de las columnas B y C
        valor_b = row[0].value
        valor_c = row[1].value
        
        # Concatenar los valores de B y C y almacenar el resultado en la columna B
        concatenated_value = f"{valor_b} {valor_c}"
        row[0].value = concatenated_value

    # Columna a eliminar: C, una vez concatenados los nombres con los apellidos
    columna_eliminar = 3

    # Elimina la columna C (Apellidos) 
    ws.delete_cols(columna_eliminar, 1)  


'''------Eliminar fecha y el resto de columnas innecesarias------'''
def delete_columns(ws):
    #Columnas a eliminar: (C, F, G, H y I) Fecha, Sexo, localidad, número de celular y TCVA
    columnas_eliminar = [3, 5, 6, 7, 8, 9]
    # Eliminar las columnas
    for col_idx in sorted(columnas_eliminar, reverse=True):
        # Iterar en orden inverso para evitar problemas con los índices cambiando
        col_letra = get_column_letter(col_idx)
        ws.delete_cols(col_idx)

    # Columna a eliminar: C, nuevamente para no dejar espacios vacíos
    columna_eliminar = 3

    # Elimina la columna C (vacía)
    ws.delete_cols(columna_eliminar, 1)  



'''------Mover datos a D5------'''
def move_data_to_D5(ws):
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['H'].width = 45

    cedula_copiar = []
    profesional_copiar = []
    hogares_pymes_cuidadoras_copiar = []

    # Iterar sobre las filas de la Hoja1
    for fila in ws.iter_rows(min_row=5, values_only=True): # Se empieza desde la celda A5
        cedula = fila[1 - 1] # columna A
        profesional = fila[2 - 1] # columna B
        hogares_pymes_cuidadoras = fila[3 - 1] # columna C

        cedula_copiar.append(cedula)
        profesional_copiar.append(profesional)
        hogares_pymes_cuidadoras_copiar.append(hogares_pymes_cuidadoras) 

        if cedula is None:
            break

    # Pegar los datos en las celdas D, E y F.
    for i, (cedula, profesional, hogares_pymes_cuidadoras) in enumerate(zip(cedula_copiar, profesional_copiar, hogares_pymes_cuidadoras_copiar), start=5):
        ws[f'D{i}'] = cedula
        ws[f'E{i}'] = profesional
        ws[f'F{i}'] = hogares_pymes_cuidadoras

    # Eliminar contenido de la columna A
    for fila in ws.iter_rows():
        celda = fila[0]  # Celda en la columna A
        celda.value = None  # Eliminar valor de la celda

    # Eliminar contenido de la columna B 
    for fila in ws.iter_rows():
        celda = fila[1]
        celda.value = None

    # Eliminar contenido de la columna C
    for fila in ws.iter_rows():
        celda = fila[2]
        celda.value = None


'''------Sección para buscar coincidencias de la hoja INFORME SOLICITUDES------'''
def encontrar_y_mover_coincidencias_cedulas_y_nombres(ws,ws2):
    # Iterar sobre las filas de la hoja de informe, comenzando desde la fila 2 (excluyendo el encabezado)
    for fila_informe in ws2.iter_rows(min_row=2, max_col=ws2.max_column):
        numero_documento_informe = fila_informe[9].value  # Columna J (índice 9)

        # Iterar sobre las filas de la hoja Hoja1, comenzando desde la fila 5 (excluyendo el encabezado)
        for fila_hoja1 in ws.iter_rows(min_row=5, max_col=ws.max_column):
            numero_documento_hoja1 = fila_hoja1[3].value  # Columna D (índice 3)

            # Si se encuentra una coincidencia, copiar los datos a la hoja de informe
            if numero_documento_informe == numero_documento_hoja1:
                fila_informe[13].value = fila_hoja1[3].value  # Columna M (índice 12)
                fila_informe[14].value = fila_hoja1[4].value  # Columna N (índice 13)


'''------Sección para buscar coincidencias de la hoja Hoja1------'''
def encontrar_y_mover_coincidencias_nombres(ws,ws2):

    # Iterar sobre las filas de la hoja Hoja1, comenzando desde la fila 5 
    for fila_hoja1 in ws.iter_rows(min_row=5, min_col=1, max_col=10):
        numero_documento_hoja1 = fila_hoja1[3].value  # Columna D (índice 3)

        # Iterar sobre las filas de la hoja INFORME SOLCITUDES, comenzando desde la fila 2 
        for fila_informe in ws2.iter_rows(min_row=2, max_col=ws2.max_column):
            numero_documento_informe = fila_informe[9].value  # Columna J (índice 9)

            # Si se encuentra una coincidencia, copiar los datos a la hoja de informe
            if numero_documento_hoja1 == numero_documento_informe:
                fila_hoja1[7].value = fila_informe[10].value  


'''------Sección para copiar  y pegar todas las expertas que NO tienen servicio------'''
def no_service_copypaste(ws,ws2):
    # Listas para almacenar cédulas y nombres completos
    cedulas_sin_servicio = []
    nombres_sin_servicio = []
    tipo_sin_servicio = []

    # Iterar sobre las filas de la Hoja1
    for fila in ws2.iter_rows(min_row=2, values_only=True):  
        cedula = fila[4 - 1]  # Columna D, Python usa base cero
        nombre = fila[5 - 1]  # Columna E
        tipo = fila[6 - 1]  #Columna F 
        servicio = fila[8 - 1]  # Columna H


        #Si el servicio (columna H) está vacío se almacenan los datos en las tuplas anteriormente definidas 
        if servicio is None or servicio.strip() == "":
            cedulas_sin_servicio.append(cedula)
            nombres_sin_servicio.append(nombre)
            tipo_sin_servicio.append(tipo)

    # Encontrar la última fila de la tabla existente (suponiendo que la columna L siempre tiene datos)
    last_row = ws.max_row

    # Calcular la fila inicial para la copia (15 filas después)
    start_row = last_row + 10

    # Pegar los datos en la Hoja1, columnas M, N y O
    for i, (cedula, nombre, tipo) in enumerate(zip(cedulas_sin_servicio, nombres_sin_servicio, tipo_sin_servicio), start=1):
        ws[f'N{start_row + i - 1}'] = cedula
        ws[f'O{start_row + i - 1}'] = nombre
        ws[f'P{start_row + i - 1}'] = tipo



'''------Función para encontrar la tabla y moverla a la celda A5 (solo para archivos .xlsx)------'''
def find_table_and_move_to_A5_xlsx(ws):

    # Encontrar la tabla en la hoja 'Hoja1'
    inicio_fila = None
    inicio_columna = None
    for fila in ws.iter_rows(min_row=1, max_col=ws.max_column):
        for cell in fila:
            if cell.value is not None:
                inicio_fila = cell.row
                inicio_columna = cell.column
                break
        if inicio_fila:        
            break
    
    # Si no se encontró la tabla, no hacer nada
    if inicio_fila is None or inicio_columna is None:
        print("No se encontró la tabla en la hoja 'Hoja1'.")
        return
    
    # Obtener los datos de la tabla
    datos_tabla = []
    for fila in ws.iter_rows(min_row=inicio_fila, min_col=inicio_columna, max_col=inicio_columna + 10):
        fila_datos = [cell.value for cell in fila]
        if all(cell is None for cell in fila_datos):
            break
        datos_tabla.append(fila_datos)
    
    # Limpiar la tabla existente
    for fila in ws.iter_rows(min_row=inicio_fila, min_col=inicio_columna, max_col=inicio_columna + 10):
        for cell in fila:
            cell.value = None
    
    # Colocar los datos a partir de A5
    for i, fila_datos in enumerate(datos_tabla):
        for j, valor in enumerate(fila_datos):
            ws.cell(row=5+i, column=1+j).value = valor


'''------Función para encontrar la tabla y moverla a la celda A5 (solo para archivos .xls)------'''
def find_table_and_move_to_A5_xls(file_path, temp):
    # Abrir el archivo .xls en modo lectura
    wb_rd = xlrd.open_workbook(file_path, formatting_info=True)
    ws_rd = wb_rd.sheet_by_name('Hoja1')

    # Encontrar la tabla en la hoja 'Hoja1'
    inicio_fila = None
    inicio_columna = None

    for row_idx in range(ws_rd.nrows):
        for col_idx in range(ws_rd.ncols):
            if ws_rd.cell_value(row_idx, col_idx) not in (None, ''):
                inicio_fila = row_idx
                inicio_columna = col_idx
                break
        if inicio_fila is not None:
            break

    # Si no se encontró la tabla, no hacer nada
    if inicio_fila is None or inicio_columna is None:
        print("No se encontró la tabla en la hoja 'Hoja1'.")
        return

    # Obtener los datos de la tabla
    datos_tabla = []
    for row_idx in range(inicio_fila, ws_rd.nrows):
        fila_datos = ws_rd.row_values(row_idx, start_colx=inicio_columna, end_colx=inicio_columna + 11)
        if all(cell in (None, '') for cell in fila_datos):
            break
        datos_tabla.append(fila_datos)

    # Crear una copia del archivo .xls para modificarlo
    wb_wr = copy(wb_rd)
    ws_wr = wb_wr.get_sheet('Hoja1')

    # Limpiar la tabla existente
    for row_idx in range(inicio_fila, inicio_fila + len(datos_tabla)):
        for col_idx in range(inicio_columna, inicio_columna + 11):
            ws_wr.write(row_idx, col_idx, '')

    # Colocar los datos a partir de A5
    for i, fila_datos in enumerate(datos_tabla):
        for j, valor in enumerate(fila_datos):
            ws_wr.write(4 + i, j, valor)

    # Guardar los cambios en el archivo .xls
    wb_wr.save(temp)



'''------Función para organizar la tabla alfabéticamente de la hoja INFORME SOLICITUDES, usando la columna N como índice o base del ordenamiento------'''
def organizar_tabla_alfabeticamente(ws2):
    # Obtener el estilo de fecha predeterminado (formato DD/MM/YYYY)
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
   
    # Obtener todos los datos de la hoja
    data = []
    columna_d = []  # Lista para almacenar los datos de la columna D
    for row in ws2.iter_rows(min_row=2, values_only=True):
        # Separar la columna D de los datos a ordenar
        columna_d.append(row[3])
        data.append([row[0], row[1], row[2]] + list(row[4:]))

    # Ordenar los datos por la columna O (índice 13 en data, índice 14 en la hoja), colocando los valores None al final
    data.sort(key=lambda x: (x[13] is None, x[13]))

    # Limpiar la hoja y escribir los datos ordenados a partir de la fila 2
    ws2.delete_rows(2, ws2.max_row)
    for i, row in enumerate(data, start=2):  # Comenzamos el índice en 2
        ws2.cell(row=i, column=1, value=row[0])
        ws2.cell(row=i, column=2, value=row[1])
        ws2.cell(row=i, column=3, value=row[2])
        ws2.cell(row=i, column=4, value=columna_d[i-2]).number_format = date_style.number_format  # Reinsertar la columna D con formato de fecha
        ws2.cell(row=i, column=5, value=row[3])
        ws2.cell(row=i, column=6, value=row[4])
        ws2.cell(row=i, column=7, value=row[5])
        ws2.cell(row=i, column=8, value=row[6])
        ws2.cell(row=i, column=9, value=row[7])
        ws2.cell(row=i, column=10, value=row[8])
        ws2.cell(row=i, column=11, value=row[9]) 
        ws2.cell(row=i, column=12, value=row[10])
        ws2.cell(row=i, column=13, value=row[11])
        ws2.cell(row=i, column=14, value=row[12])
        ws2.cell(row=i, column=15, value=row[13])
        ws2.cell(row=i, column=16, value=row[14])
    

'''------Función para organizar la tabla alfabéticamente de la hoja Hoja1, usando la columna H como índice o base del ordenamiento------'''
def organizar_tabla_alfabeticamente_hoja1(ws):
    # Obtener todos los datos de la hoja
    data = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        data.append(list(row))

    # Ordenar los datos por la columna H (índice 7 en data, índice 8 en la hoja)
    data.sort(key=lambda x: x[7] if x[7] is not None else '')

    # Limpiar la hoja y escribir los datos ordenados a partir de la fila 5
    ws.delete_rows(5, ws.max_row)
    for i, row in enumerate(data, start=5):  # en el indice 5
        ws.cell(row=i, column=4, value=row[3])
        ws.cell(row=i, column=5, value=row[4])
        ws.cell(row=i, column=6, value=row[5])
        ws.cell(row=i, column=8, value=row[7])
        

'''------Función que globaliza todas las funciones anteriormente definidas (LUNES-JUEVES)-------'''
def ejecucion_funciones2(ws,ws2):
    concatenar_nombres_apellidos(ws)
    delete_columns(ws)
    move_data_to_D5(ws)
    encontrar_y_mover_coincidencias_cedulas_y_nombres(ws,ws2)
    organizar_tabla_alfabeticamente(ws2)
    encontrar_y_mover_coincidencias_nombres(ws,ws2)
    no_service_copypaste(ws2,ws) #argumentos de hojas invertidos para mayor comodidad (originalmente ws es INFORME SOLICITUDES y ws2 es Hoja1)
    novedades_expertas(ws2)
    organizar_tabla_alfabeticamente_hoja1(ws)



