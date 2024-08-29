import tkinter as tk
from tkinter import filedialog, Label
from tkinter import messagebox
import os
import openpyxl
from manipular_INFORME_SOLICITUDES import *
from manipular_Hoja1 import *
import xlrd
import logging
import sys
import os
import xlwt
from convert_xlsx_to_xls import *
from funciones_weekend import *

def message_success():
  # Mensaje de éxito
  exito = messagebox.showinfo("Proceso completado", """
          • Cambios en la hoja "INFORME SOLICITUDES":
          - Se eliminaron las filas 1, 2, 3, y 4.
          - Columnas eliminadas: Total servicios, Tipo, Turno 
            partido, Jornada fija, Concepto: novedad y ausencias, 
            Concepto: novedad control empleados, CC experta 
            cambios, Experta cambio, Notificación SMS, 
            Notificación SMS cliente, SMS enviado cliente. 
          - Formato de fecha conservado DD/MM/YYYY.
          - Tamaño horizontal de las columnas adaptado.
          - Solicitud, Referencia Externa y Cédula modificada a 
            formato numérico.
          - Expertas que SI tienen novedad, resaltadas con color 
            amarillo.
          • Cambios en la hoja "Hoja1":
          - Nombres y apellidos concatendados.
          - Columnas eliminadas: Fecha, Sexo, localidad, número 
            de celular y TCVA eliminadas.
          - Datos trasladados a la celda D5.
          - BUSCARV desde Hoja1 a INFORME SOLICITUDES 
            número de documento y nombre completo en 
            las columnas M y N.
          - BUSCARV desde INFORME SOLICITUDES a Hoja1 
            nombre completo en la columna H.
          - Listado de expertas sin servicio copiado en las 
            columnas M, N y O.
          """)



def button_lunes_jueves(wb):
  # Seleccionar la hoja de trabajo
  ws = wb.active
          
  #buscar el marcador de modificación 
  marcador = ws['AZ1'] #Marcador en la celda AZ1
  modificado_antes = marcador.value == 'MODIFICADO' 

  if modificado_antes:
    # Mensaje de confirmación si el archivo ha sido modificado por la aplicación previamente  
    resultado = messagebox.askyesno('Confirmar modificación', 'Este archivo ya ha sido modificado previamente, si continúa, el contenido del archivo será distorsionado. ¿Desea continuar?')
    if not resultado:
      return

  # Función compuesta de funciones que ejecutan los cambios necesarios en la hoja INFORME SOLICITUDES
  ejecucion_funciones(ws)

  #Volver a cargar las hojas del archivo después de ejecurtar los cambios en la hoja INFORME SOLICITUDES
  ws=wb[wb.sheetnames[1]]
  ws2=wb[wb.sheetnames[0]]

  # Funciones que hacen los cambios en la Hoja1 
  ejecucion_funciones2(ws,ws2)

  #Agregar marcador de que el archivo ha sido modificado por la aplicación en la celdd AZ1        
  ws2['AZ1'] = 'MODIFICADO' 

  # Pedir al usuario la ruta y nombre para guardar el nuevo archivo
  filepath_save = filedialog.asksaveasfilename(title="Guardar archivo Excel modificado como", defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])
  if not filepath_save:
    return  # Salir si el usuario cancela el diálogo de guardado

  # Guardar el archivo modificado
  wb.save(filepath_save)
  message_success()
  abrir_excel(filepath_save)

def button_viernes_sabado(wb):
  # Seleccionar la hoja de trabajo
  ws = wb.active
    
  #buscar el marcador de modificación 
  marcador = ws['AZ1'] #Marcador en la celda AZ1
  modificado_antes = marcador.value == 'MODIFICADO' 

  if modificado_antes:
    # Mensaje de confirmación si el archivo ha sido modificado por la aplicación previamente  
    resultado = messagebox.askyesno('Confirmar modificación', 'Este archivo ya ha sido modificado previamente, si continúa, el contenido del archivo será distorsionado. ¿Desea continuar?')
    if not resultado:
      return
    
  # Función compuesta de funciones que ejecutan los cambios necesarios en la hoja INFORME SOLICITUDES
  ejecucion_funciones(ws)

  #Volver a cargar las hojas del archivo después de ejecurtar los cambios en la hoja INFORME SOLICITUDES
  ws=wb[wb.sheetnames[1]]
  ws2=wb[wb.sheetnames[0]]

  # Funciones que hacen los cambios en la Hoja1
  ejecucion_funciones2_viernes_sabado(ws,ws2)

  #Agregar marcador de que el archivo ha sido modificado por la aplicación en la celda AZ1        
  ws2['AZ1'] = 'MODIFICADO' 

  # Pedir al usuario la ruta y nombre para guardar el nuevo archivo
  filepath_save = filedialog.asksaveasfilename(title="Guardar archivo Excel modificado como", defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])
  if not filepath_save:
    return  # Salir si el usuario cancela el diálogo de guardado

  # Guardar el archivo modificado
  wb.save(filepath_save)
  message_success()
  abrir_excel(filepath_save)