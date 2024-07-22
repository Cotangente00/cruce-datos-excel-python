import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import os
import openpyxl
from manipular_INFORME_SOLICITUDES import *
from manipular_Hoja1 import *
from no_service import *
import xlrd

def procesar_INFORME_SOLICITUDES():
    filepath = filedialog.askopenfilename(title="Selecciona el archivo Excel a modificar", filetypes=[("Archivos Excel", "*.xlsx;*.xls")])
    if not filepath:
        return

    try:
        if filepath.endswith('.xlsx'):
            # Cargar el libro de Excel (xlsx)
            wb = openpyxl.load_workbook(filepath)
        elif filepath.endswith('.xls'):
            # Cargar el libro de Excel (xls)
            xls_workbook = xlrd.open_workbook(filepath)
            wb = openpyxl.Workbook()

            # Copiar hojas de xls a xlsx
            for sheet in xls_workbook.sheets():
                ws_xls = xls_workbook.sheet_by_name(sheet.name)
                ws_new = wb.create_sheet(sheet.name)

                for row in range(ws_xls.nrows):
                    for col in range(ws_xls.ncols):
                        cell_value = ws_xls.cell_value(row, col)
                        ws_new.cell(row=row+1, column=col+1, value=cell_value)

            # Eliminar la hoja predeterminada de Workbook
            wb.remove(wb.active)

        # Seleccionar la hoja de trabajo
        ws = wb.active

        # Aplicar las modificaciones utilizando Openpyxl
        delete_filas(ws)
        delete_ciudades_columnas(ws)
        date_format(ws)
        styles_columnSize(ws)
        int_format(ws)
        novedades_expertas(ws)
         # Crea una nueva hoja llamada "Hoja1"
        hoja_nueva = wb.create_sheet("Hoja1")

        hoja_nueva['A1'] = 'Pegar datos de expertas en la celda A5'

        # Pedir al usuario la ruta y nombre para guardar el nuevo archivo
        filepath_save = filedialog.asksaveasfilename(title="Guardar archivo Excel modificado como", defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])
        if not filepath_save:
            return  # Salir si el usuario cancela el diálogo de guardado

        # Guardar el archivo modificado
        wb.save(filepath_save)
        messagebox.showinfo("Proceso completado", "Modificaciones principales aplicadas: Columnas inncesarias eliminadas, las ciudades de: Bogotá, Cota, Chía y Cajicá filtradas exitosamente y las expertas que SI tienen novedades han sido resaltadas con color amarillo.")
        abrir_excel(filepath_save)
    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error al procesar el archivo:\n{str(e)}")


def procesar_Hoja1():
    # Abrir el archivo Excel seleccionado
    filepath = filedialog.askopenfilename(title="Selecciona el archivo Excel a modificar", filetypes=[("Archivos Excel", "*.xlsx")])
    if not filepath:
        return

    try:
        # Cargar el libro de Excel
        wb = openpyxl.load_workbook(filepath)
        # Seleccionar la hoja de trabajo
        ws = wb['Hoja1']
        

        # Aplicar las modificaciones utilizando Openpyxl
        # Aquí se coloca la lógica para modificar los datos del archivo Excel
        concatenar_nombres_apellidos(ws)
        delete_columns(ws)   
        move_data_to_D5(ws)     

        # Reescribir o guardar los cambios en el mismo archivo modificado 
        wb.save(filepath)
        messagebox.showinfo("Proceso completado", "Modificaciones principales aplicadas: Columnas innecesarias eliminadas, nombres y apellidos concatenados exitosamente y listado de expertas trasladados a la celda D5.")
        abrir_excel(filepath)
    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error al procesar el archivo:\n{str(e)}")


def no_service():
    # Abrir el archivo Excel seleccionado
    filepath = filedialog.askopenfilename(title="Selecciona el archivo Excel a modificar", filetypes=[("Archivos Excel", "*.xlsx")])
    if not filepath:
        return

    try:
        # Cargar el libro de Excel
        wb = openpyxl.load_workbook(filepath)
        # Seleccionar las hoja de trabajo
        ws = wb['INFORME SOLICITUDES']
        ws2 = wb['Hoja1']

        # Aplicar las modificaciones utilizando Openpyxl
        # Aquí se coloca la lógica para modificar los datos del archivo Excel
        no_service_copypaste(ws, ws2) 
        ws['Q2'] = 'Expertas que NO tienen servicio'  
    

        # Reescribir o guardar los cambios en el mismo archivo modificado 
        wb.save(filepath)
        messagebox.showinfo("Proceso completado", "Modificaciones principales aplicadas: listado de expertas que NO tienen servicio ha sido copiado y pegado en las columnas Q y R")
        abrir_excel(filepath)
    except Exception as e:  
        messagebox.showerror("Error", f"Ha ocurrido un error al procesar el archivo:\n{str(e)}")


# Configurar la interfaz gráfica
root = tk.Tk()
root.wm_title("Manualidades de Excel fuera")
root.geometry('320x240')
icon_path = os.path.join(os.path.dirname(__file__), 'icon.ico')
root.iconbitmap(icon_path)

btn_procesar_informe_solicitudes = tk.Button(root, text="1. Procesar INFORME SOLICITUDES", command=procesar_INFORME_SOLICITUDES)
btn_procesar_informe_solicitudes.pack(pady=20)

btn_procesar_hoja1 = tk.Button(root, text="2. Procesar Hoja1", command=procesar_Hoja1)
btn_procesar_hoja1.pack(pady=20)

btn_procesar_hoja1 = tk.Button(root, text="3. Copiar y pegar las que NO tienen servicio", command=no_service)
btn_procesar_hoja1.pack(pady=20)


root.mainloop() 