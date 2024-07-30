'''------Ordenar una tabla alfabéticamente, usando una columna como índice------'''

from openpyxl import load_workbook
#wb = load_workbook('Avance.xlsx')
#ws = wb['INFORME SOLICITUDES']

def a_z(ws):
    #Leer datos de la hoja (excluyendo los encabezados con min_row=2)
    datos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos.append(fila)

    #Ordenar los datos por la primera columna 
    datos_ordenados = sorted(datos, key=lambda x: x[0].lower())

    #Escribir los datos ordenados de vuelta a la hoja (comenzando desde la segunda fila por los encabezados)
    for i, fila in enumerate(datos_ordenados, star=2):
        for j, valor in enumerate(fila, star=1):
    ws.cell(row=i, column=j, value=valor)

#wb.save()